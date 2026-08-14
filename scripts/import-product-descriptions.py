#!/usr/bin/env python3
"""Import product descriptions from a filled workbook into Strapi commerce-products.

    python3 scripts/import-product-descriptions.py --dry      # report only, no writes
    python3 scripts/import-product-descriptions.py            # write to Strapi
    python3 scripts/import-product-descriptions.py --slug dell-xps-16-2026-512gb
    python3 scripts/import-product-descriptions.py --restore backups/descriptions-<ts>.json

Reads the "Products" sheet and writes its `description` column to the matching
product's `description` field.

This overwrites. Every product in the workbook already has a description, and
most of those read as written prose rather than scraped shop copy, so this is
not filling a gap — it is replacing one piece of writing with another. Two
consequences are built in:

  * Every existing description is written to backups/ before anything changes,
    and `--restore` puts them back. A bulk overwrite of editorial content that
    cannot be undone is not worth the few lines it saves.
  * The workbook's `sources` and `confidence` columns are stored alongside the
    text in `specs.descriptionSource`. A description whose provenance is lost
    cannot be checked later, and these were written from cited pages.

The two can disagree on facts — the Dell XPS 16 2026 rows are the known case,
where the stored text says Panther Lake / Core Ultra Series 3 and the workbook
says Core Ultra Series 2 — so the report prints what changed rather than only
how many rows moved.

Not imported: the `specifications` column. That is a separate field and a
separate decision from the descriptions this was asked to bring in.

Env (.env.local): STRAPI_INTERNAL_URL or NEXT_PUBLIC_STRAPI_URL, plus
STRAPI_WRITE_TOKEN if set, else STRAPI_API_TOKEN.
"""

import argparse
import json
import os
import re
import sys
import urllib.error
import urllib.parse
import urllib.request
from datetime import date, datetime, timezone
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
DEFAULT_WORKBOOK = "/opt/assets/nxt-bargains-other-categories-with-descriptions.xlsx"
SHEET = "Products"


def load_env():
    env_file = ROOT / ".env.local"
    if not env_file.exists():
        return
    for line in env_file.read_text().splitlines():
        m = re.match(r"^\s*([A-Z0-9_]+)\s*=\s*(.*?)\s*$", line)
        if m and m.group(1) not in os.environ:
            os.environ[m.group(1)] = m.group(2)


def read_rows(workbook_path):
    wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    if SHEET not in wb.sheetnames:
        sys.exit(f"{workbook_path}: no '{SHEET}' sheet")
    rows = wb[SHEET].iter_rows(values_only=True)
    header = [str(c).strip() if c else "" for c in next(rows)]
    idx = {name: i for i, name in enumerate(header)}
    for column in ("slug", "description"):
        if column not in idx:
            sys.exit(f"{workbook_path}: '{SHEET}' is missing the '{column}' column")
    out = []
    for row in rows:
        if not row or not row[idx["slug"]]:
            continue
        out.append({key: row[i] for key, i in idx.items()})
    wb.close()
    return out


class Strapi:
    def __init__(self, base, token, timeout=60):
        self.base = base.rstrip("/")
        self.token = token
        self.timeout = timeout

    def _request(self, method, path, payload=None):
        url = f"{self.base}/api/{path}"
        body = json.dumps(payload).encode() if payload is not None else None
        req = urllib.request.Request(url, data=body, method=method)
        req.add_header("Authorization", f"Bearer {self.token}")
        req.add_header("Content-Type", "application/json")
        try:
            with urllib.request.urlopen(req, timeout=self.timeout) as res:
                return json.load(res)
        except urllib.error.HTTPError as err:
            detail = err.read().decode(errors="replace")[:400]
            raise SystemExit(f"Strapi {err.code} on {method} {url}: {detail}")

    def get_product(self, slug):
        query = urllib.parse.urlencode({
            "filters[slug][$eq]": slug,
            "fields[0]": "slug",
            "fields[1]": "name",
            "fields[2]": "description",
            "fields[3]": "specs",
        })
        data = self._request("GET", f"commerce-products?{query}").get("data") or []
        return data[0] if data else None

    def update(self, document_id, payload):
        return self._request("PUT", f"commerce-products/{document_id}", {"data": payload})


def restore(strapi, path):
    entries = json.loads(Path(path).read_text())["descriptions"]
    print(f"Restoring {len(entries)} description(s) from {path}\n")
    for entry in entries:
        product = strapi.get_product(entry["slug"])
        if not product:
            print(f"  !  {entry['slug']}: not in Strapi")
            continue
        strapi.update(product["documentId"], {"description": entry["description"]})
        print(f"  ↩  {entry['slug']}")
    print(f"\nRestored: {len(entries)}")


def main():
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--workbook", default=DEFAULT_WORKBOOK, help=f"xlsx to import (default: {DEFAULT_WORKBOOK})")
    ap.add_argument("--dry", action="store_true", help="report what would change, write nothing")
    ap.add_argument("--slug", action="append", help="only this slug (repeatable)")
    ap.add_argument("--limit", type=int, help="stop after N products")
    ap.add_argument("--min-confidence", choices=["low", "medium", "high"], default="low",
                    help="skip rows below this confidence (default: low, i.e. import everything)")
    ap.add_argument("--restore", metavar="BACKUP", help="put the descriptions in a backup file back, then exit")
    ap.add_argument("--report", help="write a JSON report of the run to this path")
    args = ap.parse_args()

    load_env()
    base = os.environ.get("STRAPI_INTERNAL_URL") or os.environ.get("NEXT_PUBLIC_STRAPI_URL")
    token = os.environ.get("STRAPI_WRITE_TOKEN") or os.environ.get("STRAPI_API_TOKEN")
    if not base or not token:
        sys.exit("Set STRAPI_INTERNAL_URL/NEXT_PUBLIC_STRAPI_URL and STRAPI_API_TOKEN in .env.local")
    strapi = Strapi(base, token)

    if args.restore:
        restore(strapi, args.restore)
        return

    rank = {"low": 0, "medium": 1, "high": 2}
    rows = read_rows(args.workbook)
    if args.slug:
        wanted = set(args.slug)
        rows = [r for r in rows if r["slug"] in wanted]
        for slug in sorted(wanted - {r["slug"] for r in rows}):
            print(f"  ?  {slug}: not in the workbook")

    stamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    backup_path = ROOT / "backups" / f"descriptions-{stamp}.json"
    report = {"updated": [], "unchanged": [], "skipped_confidence": [], "missing_in_strapi": [], "empty": []}
    backup = []

    print(f"{'DRY RUN — ' if args.dry else ''}{len(rows)} row(s) from {args.workbook} → {base}\n")

    for row in rows:
        slug = row["slug"]
        text = (row.get("description") or "").strip()
        confidence = (row.get("confidence") or "").strip().lower()

        if args.limit and len(report["updated"]) >= args.limit:
            break
        if not text:
            report["empty"].append(slug)
            print(f"  –  {slug}: no description in the workbook, skipped")
            continue
        if confidence in rank and rank[confidence] < rank[args.min_confidence]:
            report["skipped_confidence"].append({"slug": slug, "confidence": confidence})
            print(f"  –  {slug}: confidence '{confidence}' below --min-confidence, skipped")
            continue

        product = strapi.get_product(slug)
        if not product:
            report["missing_in_strapi"].append(slug)
            print(f"  !  {slug}: no commerce-product with this slug")
            continue

        current = (product.get("description") or "").strip()
        if current == text:
            report["unchanged"].append(slug)
            continue

        backup.append({"slug": slug, "documentId": product["documentId"], "description": product.get("description")})

        specs = product.get("specs")
        specs = dict(specs) if isinstance(specs, dict) else {}
        sources = [s.strip() for s in re.split(r"[\n,]+", str(row.get("sources") or "")) if s.strip().startswith("http")]
        specs["descriptionSource"] = {
            "source": "editorial-workbook",
            "workbook": os.path.basename(args.workbook),
            "confidence": confidence or None,
            "sources": sources,
            "importedAt": date.today().isoformat(),
        }

        if not args.dry:
            # Written before the update, and rewritten after every row, so a run
            # that dies halfway still leaves a backup covering everything it
            # actually changed.
            backup_path.parent.mkdir(parents=True, exist_ok=True)
            backup_path.write_text(json.dumps({"takenAt": stamp, "descriptions": backup}, indent=2))
            strapi.update(product["documentId"], {"description": text, "specs": specs})

        report["updated"].append({
            "slug": slug, "confidence": confidence,
            "wasChars": len(current), "nowChars": len(text), "sources": len(sources),
        })
        print(f"  {'·' if args.dry else '✓'}  {slug}: {len(current)} → {len(text)} chars ({confidence}, {len(sources)} source(s))")

    print(
        f"\n{'Would update' if args.dry else 'Updated'}: {len(report['updated'])}"
        f" · identical already: {len(report['unchanged'])}"
        f" · below confidence: {len(report['skipped_confidence'])}"
        f" · blank in workbook: {len(report['empty'])}"
        f" · not in Strapi: {len(report['missing_in_strapi'])}"
    )
    if backup and not args.dry:
        print(f"Previous descriptions saved to {backup_path}")
        print(f"  undo with: python3 scripts/import-product-descriptions.py --restore {backup_path}")
    if args.report:
        report["ranAt"] = datetime.now(timezone.utc).isoformat()
        report["dryRun"] = args.dry
        report["backup"] = str(backup_path) if backup and not args.dry else None
        Path(args.report).write_text(json.dumps(report, indent=2))
        print(f"Report: {args.report}")


if __name__ == "__main__":
    main()
